import pandas as pd
import openpyxl
from openpyxl import Workbook
from openpyxl import load_workbook
import statistics
df = pd.read_excel('/Users/cmason/Library/Containers/com.microsoft.Excel/Data/Downloads/did_test3.xlsx')
wb = load_workbook('/Users/cmason/Library/Containers/com.microsoft.Excel/Data/Downloads/did_test2.xlsx')
ws = wb.active
class Mouse:
    def __init__(self, **kwargs):
        self.__dict__.update(kwargs)

    @staticmethod
    def data_dict():
        empty_dict = dict.fromkeys(['NUMBER', 'WEIGHT', 'D1VI', 'D2VI', 'D3VI', 'D4VI', 'D1VF', 'D2VF', 'D3VF', 'D4VF',
                                    'D1DC1I', 'D2DC1I', 'D3DC1I', 'D4DC1I', 'D1DC1F', 'D2DC1F', 'D3DC1F', 'D4DC1F',
                                    'D1DC2I', 'D2DC2I', 'D3DC2I', 'D4DC2I', 'D1DC2F', 'D2DC2F', 'D3DC2F', 'D4DC2F'])
        print("The Mouse House of DID(TM) only accepts numbers. Any characters or white space will assume the value "
              "'None' on data sheet")

        for key in empty_dict.keys():
            while True:
                try:
                    empty_dict[key] = float(input("What is the value of " + key + "?"))
                except ValueError:
                    print('Please put a valid number.')
                    continue
                else:
                    break
        mouse_data = empty_dict
        return mouse_data





            # try:
            #     empty_dict[key] = float(input("What is the value of " + key + "?"))
            # except isinstance(empty_dict[key], float):
            #     print('please try again')
            #     empty_dict[key] = float(input("What is the value of " + key + "?"))


    @staticmethod
    def diff_tuple_list():     #x

        #make variable of a list
        data = test_dict   #Mouse.data_dict()


        #stuff the value differences into the list

        value = data.get('b') - data.get('a')
        print(value)

        #return your list variable


test_dict = {'aghf':1,'b':2,'c':3,'d':4,'e':5,'f':6,'g':7,'h':8,'i':9,'j':10,'k':11,'l':12,'m':13,
           'n':14,'o':15,'p':16,'q':17,'r':18, 's':19,'t':20,'u':21,'v':22,'w':23,'x':24,'y':25,'z':26}
# # y ={'mouse_number': 1, 'wgt': 2.0, 'D1VI': 3.0, 'D2VI': 4.0, 'D3VI': 5.0, 'D4VI': 6.0, 'D1VF': 7.0, 'D2VF': 8.0,
#      'D3VF': 9.0, 'D4VF': 10.0, 'D1DC1I':11.0, 'D2DC1I': 12.0, 'D3DC1I': 13.0, 'D4DC1I': 14.0, 'D1DC1F': 15.0,
#      'D2DC1F': 16.0, 'D3DC1F': 17.0, 'D4DC1F': 18.0, 'D1DC2I': 19.0, 'D2DC2I': 20.0, 'D3DC2I': 21.0, 'D4DC2I': 22.0,
#      'D1DC2F': 23.0, 'D2DC2F': 24.0, 'D3DC2F': 25.0, 'D4DC2F': 26.0}


choice = 0
#TODO: read saved mouse data
while choice != -1:
    print("Hello and welcome to the Mouse House of DID(TM). Here we will keep the data of our mice."
          "\nEnter the correspoding "
          "number to the task that you would like to do and enter -1 to quit this program")
    print("1- Would you like to enter data on a new mouse?")
    # print("2- Would you like to delete a mouse?") #by number so conditional check on number
    # print("3- Would you like to export the Mouse House of Data(TM) to a .csv?")
    # print("4- Would you like to enter Nightmare mode?")

    # print("5- View the tennants of the current Mouse House?") #The Mouse House of Data(TM) happily homes these mice:
    choice = int(input("What option would you like? "))
    print(" ")

    if choice == 1:

        # ws.append(test_dict)
        mr_mouse = Mouse.data_dict()
        list1 = list(mr_mouse.values())

        # list 2 stuff
        diff_v_day1 = mr_mouse.get('D1VF') - mr_mouse.get('D1VI')
        diff_v_day2 = mr_mouse.get('D3VF') - mr_mouse.get('D3VI')
        diff_v_day3 = mr_mouse.get('D1VF') - mr_mouse.get('D1VI')
        diff_v_day4 = mr_mouse.get('D4VF') - mr_mouse.get('D4VI')
        diff_dc1_day1 = mr_mouse.get('D1DC1F') - mr_mouse.get('D1DC1I')
        diff_dc1_day2 = mr_mouse.get('D2DC1F') - mr_mouse.get('D2DC1I')
        diff_dc1_day3 = mr_mouse.get('D3DC1F') - mr_mouse.get('D3DC1I')
        diff_dc1_day4 = mr_mouse.get('D4DC1F') - mr_mouse.get('D4DC1I')
        diff_dc2_day1 = mr_mouse.get('D1DC2F') - mr_mouse.get('D1DC2I')
        diff_dc2_day2 = mr_mouse.get('D2DC2F') - mr_mouse.get('D2DC2I')
        diff_dc2_day3 = mr_mouse.get('D3DC2F') - mr_mouse.get('D3DC2I')
        diff_dc2_day4 = mr_mouse.get('D4DC2F') - mr_mouse.get('D4DC2I')
        list2 = [diff_v_day1, diff_v_day2, diff_v_day3, diff_v_day4, diff_dc1_day1, diff_dc1_day2,
                 diff_dc1_day3, diff_dc1_day4, diff_dc2_day1, diff_dc2_day3, diff_dc2_day3, diff_dc2_day4]

        # list 3 stuff
        true_change_d1 = diff_v_day1 - ((diff_dc1_day1 + diff_dc2_day1)/2)
        true_change_d2 = diff_v_day2 - ((diff_dc1_day2 + diff_dc2_day2)/2)
        true_change_d3 = diff_v_day3 - ((diff_dc1_day3 + diff_dc2_day3)/2)
        true_change_d4 = diff_v_day4 - ((diff_dc1_day4 + diff_dc2_day4)/2)
        avg_day_change = (diff_v_day1 + diff_v_day2 + diff_v_day3 + diff_v_day4)/4
        avg_day_change_wgt = avg_day_change/mr_mouse.get('WEIGHT')
        avg_dc_across_days = (diff_dc1_day1 + diff_dc2_day2 + diff_dc1_day3 + diff_dc1_day4 + diff_dc2_day1 +
                              diff_dc2_day2 + diff_dc2_day3 + diff_dc2_day4)/8
        avg_day_change_wgt_and_dc = (avg_day_change - avg_dc_across_days)/mr_mouse.get('WEIGHT')

        list3 = [true_change_d1, true_change_d2, true_change_d3, true_change_d4, avg_day_change_wgt,
                 avg_day_change_wgt_and_dc]

        z = list1 + list2 + list3
        ws.append(z)
        wb.save('/Users/cmason/Library/Containers/com.microsoft.Excel/Data/Downloads/did_test2.xlsx')

        # dr_mouse = list(mr_mouse.keys())
    else:
        print("pick a number")

    # if choice == 2:
    # if choice == 3:
    # if choice == 4:
    # if choice == 5:
#TODO: api that saves mouse house to master_csv
#TODO: api that reads master_csv and instantiates a mouse house.

print(list(ws.iter_rows()))
